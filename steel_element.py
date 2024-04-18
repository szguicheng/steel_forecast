import requests
import openpyxl
from lxml import etree
from bs4 import BeautifulSoup


def extract_max_values(url):
    try:
        headers = {
            'User-Agent':
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3',
            'Referer': 'https://www.caishuku.com/',
        }
        response = requests.get(url, headers=headers)
        print(url, response.status_code)
    except requests.exceptions.RequestException as e:
        print("An error occurred:", e)
    soup = BeautifulSoup(response.text, 'html.parser')

    max_value_td = soup.find('td', text='最大值')  # 找到包含 "最大值" 的 td 元素
    tr = max_value_td.parent  # 找到该 td 元素的父元素，即 tr 元素
    max_values = [td.text for td in tr.find_all('td')]  # 提取 tr 元素中所有 td 元素的文本
    max_values_cor = [value if value != '-' else '0' for value in max_values]

    return max_values_cor


def extract_min_values(url):
    try:
        headers = {
            'User-Agent':
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3',
            'Referer': 'https://www.caishuku.com/',
        }
        response = requests.get(url, headers=headers)
        print(url, response.status_code)
    except requests.exceptions.RequestException as e:
        print("An error occurred:", e)
    soup = BeautifulSoup(response.text, 'html.parser')

    min_value_td = soup.find('td', text='最小值')  # 找到包含 "最小值" 的 td 元素
    tr = min_value_td.parent  # 找到该 td 元素的父元素，即 tr 元素
    min_values = [td.text for td in tr.find_all('td')]  # 提取 tr 元素中所有 td 元素的文本
    min_values_cor = [value if value != '-' else '0' for value in min_values]
    return min_values_cor


def extract_values(url):
    try:
        headers = {
            'User-Agent':
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3',
            'Referer': 'https://www.caishuku.com/',
        }
        response = requests.get(url, headers=headers)
        print(url, response.status_code)
    except requests.exceptions.RequestException as e:
        print("An error occurred:", e)
    soup = BeautifulSoup(response.text, 'html.parser')

    value_td = soup.find('th', text='成分')
    tr = value_td.parent
    values = [th.text for th in tr.find_all('th')]

    return values


def steel_name(url):
    try:
        headers = {
            'User-Agent':
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3',
            'Referer': 'https://www.caishuku.com/',
        }
        response = requests.get(url, headers=headers)
        print(url, response.status_code)
    except requests.exceptions.RequestException as e:
        print("An error occurred:", e)
    soup = BeautifulSoup(response.text, 'html.parser')

    def is_steel_name(text):
        return text.strip() == '牌号'

    steel_name_td = soup.find('td',
                              text=is_steel_name).find_next_sibling().text
    return steel_name_td


url = 'https://www.caishuku.com/material/detail.php?mid=azhlNWIxMTczNTQ='
max_values = extract_max_values(url)
min_values = extract_min_values(url)
values = extract_values(url)

name = steel_name(url)
# Save max_values and min_values to an Excel file
wb = openpyxl.Workbook()
sheet = wb.active

# Write max_values to the first row
for i, value in enumerate(max_values):
    sheet.cell(row=2, column=2 * (i + 1), value=value)

# Write min_values to the second row
for i, value in enumerate(min_values):
    sheet.cell(row=2, column=2 * (i + 1) - 1, value=value)

for i, value in enumerate(values):
    sheet.cell(row=1, column=2 * (i + 1) - 1, value=value)

# Save the Excel file
wb.save(
    f'/Users/guicheng/Documents/大学/本科学业/毕业设计/腐蚀数据/steel_elements/{name}元素含量.xlsx'
)
print(f'{name}的元素含量保存成功！')
